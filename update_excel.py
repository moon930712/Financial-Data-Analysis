import os
import glob
import psycopg2
import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def load_env():
    with open(r'c:\Users\Hubnet\antigravity\.env', 'r') as f:
        for line in f:
            line = line.strip()
            if '= ' in line or '=' in line:
                try:
                    k, v = line.split('=', 1)
                    os.environ[k.strip()] = v.strip()
                except: pass

load_env()

conn = psycopg2.connect(
    host=os.environ.get('DB_HOST'),
    port=os.environ.get('DB_PORT', 5432),
    dbname=os.environ.get('DB_NAME'),
    user=os.environ.get('DB_USER'),
    password=os.environ.get('DB_PASSWORD')
)

excel_path = r'c:\Users\Hubnet\antigravity\excel\Finex_프로젝트정리_v0.3_문기성2.xlsx'

target_date = '2026-03-18'

df_summary_sheet = pd.read_excel(excel_path, sheet_name='MACD 적용 수기 검증결과')
df_detail_sheet = pd.read_excel(excel_path, sheet_name='MACD 적용 수기 검증결과 (종목)')

summary_new_rows = []
detail_new_rows = []

strategy_map = {}
for _, row in df_summary_sheet.iterrows():
    if not pd.isna(row.get('투자전략(상세)')):
        strategy_map[row['투자전략(상세)']] = (row.get('구분'), row.get('투자전략'))

# Process "수익률" dir
profit_dir = r'C:\Users\Hubnet\antigravity\excel\수익률'
if os.path.exists(profit_dir):
    for filepath in glob.glob(os.path.join(profit_dir, '*.txt')):
        filename = os.path.basename(filepath)
        strategy_detail = filename.replace('_수익률.txt', '').replace('.txt', '')
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                query = f.read()
                
            df = pd.read_sql_query(query, conn)
            date_col = next((c for c in df.columns if c in ['일자', 'date', '추천일자']), None)
            if not date_col: continue
            
            df[date_col] = df[date_col].astype(str).str[:10]
            df_filtered = df[df[date_col] == target_date].copy()
            
            if df_filtered.empty: continue
                
            gubun, strategy = strategy_map.get(strategy_detail, ('단기투자전략', '종가매매'))
            
            for _, row in df_filtered.iterrows():
                new_row = {
                    '구분': gubun, '투자전략': strategy, '투자전략(상세)': strategy_detail,
                    '일자': target_date, '종목수': row.get('종목수', row.get('item_cnt', 0))
                }
                for i in range(1, 11):
                    new_row[f'D+{i} 수익률(%)'] = row.get(f'd{i}_수익률')
                    new_row[f'D+{i} 승률(%)'] = row.get(f'd{i}_승률')
                    new_row[f'D+{i} 승무패'] = row.get(f'd{i}_승무패')
                summary_new_rows.append(new_row)
        except Exception as e:
            print(f"Error in {filename}: {e}")

# Process "종목" dir
stock_dir = r'C:\Users\Hubnet\antigravity\excel\종목'
if os.path.exists(stock_dir):
    for filepath in glob.glob(os.path.join(stock_dir, '*.txt')):
        filename = os.path.basename(filepath)
        strategy_detail = filename.replace('_종목.txt', '').replace('.txt', '')
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                query = f.read()
                
            df = pd.read_sql_query(query, conn)
            date_col = next((c for c in df.columns if c in ['일자', 'date', '추천일자', 'stck_bsop_date']), None)
            if not date_col: continue
            
            df[date_col] = df[date_col].astype(str).str[:10]
            df_filtered = df[df[date_col] == target_date].copy()
            if df_filtered.empty: continue
                
            gubun, strategy = strategy_map.get(strategy_detail, ('단기투자전략', '종가매매'))
            
            for _, row in df_filtered.iterrows():
                new_row = {
                    '구분': gubun, '투자전략': strategy, '투자전략(상세)': strategy_detail,
                    '일자': target_date, '종목명': row.get('종목명', row.get('stock_name', ''))
                }
                for i in range(1, 11):
                    new_row[f'D+{i} 수익률(%)'] = row.get(f'profit_rate{i}', row.get(f'd{i}_수익률'))
                detail_new_rows.append(new_row)
        except Exception as e:
            print(f"Error in {filename}: {e}")

conn.close()

if summary_new_rows or detail_new_rows:
    print(f"Found {len(summary_new_rows)} summary rows, {len(detail_new_rows)} detail rows.")
    wb = load_workbook(excel_path)
    
    def append_sheet(sheet_name, new_rows_list):
        if not new_rows_list: return
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for r_dict in new_rows_list:
            row_data = [r_dict.get(c, '') for c in header_row]
            ws.append(row_data)

    append_sheet('MACD 적용 수기 검증결과', summary_new_rows)
    append_sheet('MACD 적용 수기 검증결과 (종목)', detail_new_rows)
    
    # Save cleanly 
    wb.save(excel_path)
    print("Excel file successfully appended and saved.")
else:
    print("NO DATA FOUND FOR DATE", target_date)
