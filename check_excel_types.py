from openpyxl import load_workbook

file_path = r'C:\Users\Hubnet\antigravity\excel\Finex_프로젝트정리_v0.3_문기성2.xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb.active # Assuming 시트1 is the active one or first one

print(f"Sheet Name: {ws.title}")

# Check a few rows from the return columns (e.g. Columns F, I, L... which correspond to D+N returns)
# Headers are likely in row 1. Data starts from row 2.
for row in ws.iter_rows(min_row=2, max_row=5, min_col=6, max_col=10):
    for cell in row:
        print(f"Cell {cell.coordinate}: Value='{cell.value}', Type={type(cell.value)}")
